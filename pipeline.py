"""
Data pipeline: parses a RingCentral Performance Report Excel export and
produces the analysis behind the AI Receptionist business-case deck.

Methodology (locked, validated against real FBM export):
- Unit of analysis = session (grouped by Session Id), inbound only.
- Session ID deduplication removes phantom ring legs (one call routed to N
  agents counts once, not N times).
- Sub-5-second filter: a session is excluded only when its LONGEST leg's Call
  Length is <= 5s. Call Length includes ring time, so genuine missed calls that
  rang 20-30s are kept; only calls where the whole thing lasted <5s (misdials,
  wrong numbers, auto-dialer hang-ups too brief to answer) are removed. This
  lowers the missed-call count, so it is conservative by construction.
- Per-session outcome priority: Answered > VM/Abandoned > VM/Missed >
  Abandoned > Missed. VM/Abandoned and VM/Missed are kept separate internally
  (never merged); the deck's "voicemail" line is a display-only grouping.
- Queue tiering (A/B/C/D) is assigned externally (Claude) and Tier D
  (back-office) queues are excluded from the headline universe.
- Business hours = Mon-Fri, 07:00-18:00 local (per Call Start Time as exported).
- Repeat callers = distinct From Number with 2+ unanswered sessions.
"""

from __future__ import annotations
import re
import numpy as np
import pandas as pd
from pathlib import Path
from dataclasses import dataclass, field
from typing import Optional
import datetime as dt

OUTCOME_ANSWERED = "Answered"
OUTCOME_VM_ABANDONED = "VM/Abandoned"
OUTCOME_VM_MISSED = "VM/Missed"
OUTCOME_ABANDONED = "Abandoned"
OUTCOME_MISSED = "Missed"

# Display grouping for the "how they were missed" slide
DISPLAY_ABANDONED = "Abandoned while ringing"
DISPLAY_VOICEMAIL = "Went to voicemail"
DISPLAY_RANG_OUT = "Rang out — no answer"

SPAM_MAX_SECONDS = 5
BUSINESS_DAYS = {0, 1, 2, 3, 4}  # Mon-Fri
BUSINESS_START_HOUR = 7
BUSINESS_END_HOUR = 18


@dataclass
class QueueStats:
    name: str
    inbound: int = 0
    answered: int = 0
    vm_abandoned: int = 0
    vm_missed: int = 0
    abandoned: int = 0
    missed: int = 0
    answered_under_60: int = 0   # answered calls with talk time < 60s
    tier: str = ""             # A / B / C / D, assigned later
    classification: str = ""   # human-readable label, assigned later

    @property
    def total_missed(self) -> int:
        return self.vm_abandoned + self.vm_missed + self.abandoned + self.missed

    @property
    def abandoned_total(self) -> int:
        """Callers who waited in queue then hung up (abandoned + abandoned-to-VM)."""
        return self.abandoned + self.vm_abandoned

    @property
    def abandon_rate(self) -> float:
        return self.abandoned_total / self.inbound if self.inbound else 0.0

    @property
    def miss_rate(self) -> float:
        return self.total_missed / self.inbound if self.inbound else 0.0

    @property
    def answer_rate(self) -> float:
        return self.answered / self.inbound if self.inbound else 0.0


@dataclass
class QueueAbandon:
    """One queue's abandonment from the RingCentral Queues performance report."""
    name: str
    ext: str = ""
    inbound: int = 0
    answered: int = 0
    abandoned: int = 0
    tier: str = ""
    classification: str = ""

    @property
    def abandon_rate(self) -> float:
        return self.abandoned / self.inbound if self.inbound else 0.0

    @property
    def answer_rate(self) -> float:
        return self.answered / self.inbound if self.inbound else 0.0


@dataclass
class QueuesReport:
    """Account-level abandonment from the RingCentral *Queues* performance report.

    This is an AGGREGATE (per-queue totals) over a different population than the
    per-call Calls export: it counts ONLY calls that entered an ACD call queue.
    It is the only source that distinguishes abandoned-in-queue callers.
    """
    inbound: int = 0
    answered: int = 0
    abandoned: int = 0
    voicemail: int = 0
    avg_speed_answer: str = ""
    avg_wait: str = ""
    longest_wait: str = ""
    sla_pct: float = 0.0
    queues: list[QueueAbandon] = field(default_factory=list)

    @property
    def abandon_rate(self) -> float:
        return self.abandoned / self.inbound if self.inbound else 0.0

    @property
    def answer_rate(self) -> float:
        return self.answered / self.inbound if self.inbound else 0.0


@dataclass
class PipelineResult:
    # Raw / dedup counts
    raw_inbound_legs: int
    inbound_sessions: int          # after dedup, before spam filter
    phantom_legs_removed: int
    spam_sessions_removed: int

    # Universe used for headline figures (A+B+C queues, spam removed)
    universe_sessions: int
    answered: int
    vm_abandoned: int
    vm_missed: int
    abandoned: int
    missed: int

    # Derived analytics
    business_hours_miss_pct: float
    repeat_callers: int
    hourly_missed: dict[int, int]  # hour(7..17) -> missed count
    days_in_period: int

    # AIR opportunity signals (Ricky Love expansion)
    answered_under_60: int = 0     # answered calls with talk time < 60s

    # Miss-rate-by-hour analysis (reference deck slide 3)
    hourly_miss_rate: dict[int, float] = field(default_factory=dict)   # hour 0..23 -> miss rate
    hourly_inbound: dict[int, int] = field(default_factory=dict)       # hour 0..23 -> inbound count
    after_hours_miss_lo: float = 0.0   # 6pm-6am miss-rate range
    after_hours_miss_hi: float = 0.0
    after_hours_miss_rate: float = 0.0 # 6pm-6am volume-weighted miss rate
    weekend_miss_lo: float = 0.0       # Sat/Sun miss-rate range
    weekend_miss_hi: float = 0.0
    weekend_miss_rate: float = 0.0     # Sat/Sun volume-weighted miss rate
    midday_miss_rate: float = 0.0      # 11a-2p miss rate

    # ROI model inputs
    avg_answered_minutes: float = 3.0  # mean talk time of answered calls (minutes)

    # Direct-dial calls (personal extensions) — excluded from the headline
    # missed-revenue universe, surfaced only as a smaller secondary figure.
    direct_total: int = 0
    direct_missed: int = 0
    direct_abandoned: int = 0

    queue_stats: dict[str, QueueStats] = field(default_factory=dict)
    queues_report: Optional[QueuesReport] = None   # real abandoned data (2nd upload)
    reporting_period: str = ""
    reconciliation_ok: bool = True
    reconciliation_note: str = ""
    sessions_df: Optional[pd.DataFrame] = None

    @property
    def total_missed(self) -> int:
        return self.vm_abandoned + self.vm_missed + self.abandoned + self.missed

    @property
    def voicemail_total(self) -> int:
        return self.vm_abandoned + self.vm_missed

    @property
    def abandoned_total(self) -> int:
        """Callers who waited in queue then hung up (abandoned + abandoned-to-VM)."""
        return self.vm_abandoned + self.abandoned

    @property
    def under_60_pct(self) -> float:
        """Share of answered calls that were under 60s of talk time."""
        return self.answered_under_60 / self.answered if self.answered else 0.0

    @property
    def miss_rate(self) -> float:
        return self.total_missed / self.universe_sessions if self.universe_sessions else 0.0

    @property
    def answer_rate(self) -> float:
        return self.answered / self.universe_sessions if self.universe_sessions else 0.0

    @property
    def misses_per_day(self) -> float:
        return self.total_missed / self.days_in_period if self.days_in_period else 0.0


def _to_seconds(val) -> float:
    if pd.isna(val):
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip()
    if ":" in s:
        parts = s.split(":")
        try:
            parts = [float(p) for p in parts]
        except ValueError:
            return 0.0
        if len(parts) == 3:
            return parts[0] * 3600 + parts[1] * 60 + parts[2]
        if len(parts) == 2:
            return parts[0] * 60 + parts[1]
    try:
        return float(s)
    except ValueError:
        return 0.0


def _parse_start_time(val):
    if pd.isna(val):
        return None
    for fmt in ("%m/%d/%Y %I:%M:%S %p", "%m/%d/%Y %H:%M:%S", "%Y-%m-%d %H:%M:%S"):
        try:
            return dt.datetime.strptime(str(val).strip(), fmt)
        except ValueError:
            continue
    try:
        return pd.to_datetime(val).to_pydatetime()
    except Exception:
        return None


def _classify_session(group: pd.DataFrame) -> str:
    results = set(group["Result"].dropna().str.strip())
    handle_times = group["_handle_seconds"]

    if (handle_times > 0).any() or "Answered" in results:
        return OUTCOME_ANSWERED
    if "VM/Abandoned" in results:
        return OUTCOME_VM_ABANDONED
    if "VM/Missed" in results:
        return OUTCOME_VM_MISSED
    if "Abandoned" in results:
        return OUTCOME_ABANDONED
    return OUTCOME_MISSED


# Columns we actually consume downstream. Reading only these (vs. the ~20+ in a
# full export) keeps peak memory far below Render's 512MB cap.
_CALLS_NEEDED_COLS = [
    "Session Id", "Call Direction", "Result", "Call Length", "Queue",
    "Handle Time", "Call Start Time", "From Number",
]
_CALLS_REQUIRED_COLS = {"Session Id", "Call Direction", "Result", "Call Length", "Queue"}


def _read_calls_sheet_streaming(path: Path) -> pd.DataFrame:
    """Stream the Calls sheet with openpyxl read_only, keeping only needed
    columns and inbound legs, to avoid loading the whole 40MB workbook at once.
    """
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        sheet_name = wb.sheetnames[0]
        for name in wb.sheetnames:
            if name.strip().lower() == "calls":
                sheet_name = name
                break
        else:
            for name in wb.sheetnames:
                if "call" in name.lower():
                    sheet_name = name
                    break
        ws = wb[sheet_name]

        rows = ws.iter_rows(values_only=True)
        try:
            header = next(rows)
        except StopIteration:
            raise ValueError("The Calls export sheet is empty.")
        header = [str(h).strip() if h is not None else "" for h in header]

        missing = _CALLS_REQUIRED_COLS - set(header)
        if missing:
            raise ValueError(
                f"Missing expected columns: {missing}. Found: {header}")

        # Map each needed column to its position; record the Call Direction index
        # so we can filter inbound while streaming.
        col_idx = {name: header.index(name) for name in _CALLS_NEEDED_COLS
                   if name in header}
        dir_i = col_idx["Call Direction"]
        # "Call Direction" is only used to filter inbound legs while streaming;
        # it's never read downstream, so don't store a 188k-row copy of it.
        keep_names = [n for n in col_idx if n != "Call Direction"]
        keep_pos = [col_idx[n] for n in keep_names]
        ncols = len(header)

        data = {n: [] for n in keep_names}
        for row in rows:
            if dir_i >= len(row):
                continue
            d = row[dir_i]
            if d is None or str(d).strip().lower() != "inbound":
                continue
            for n, p in zip(keep_names, keep_pos):
                v = row[p] if p < len(row) else None
                data[n].append("" if v is None else str(v))
    finally:
        wb.close()

    # Build the frame column-by-column, freeing each source list as it is
    # consumed. pd.DataFrame(data) would keep BOTH the Python lists and the
    # copied frame resident at once — a transient ~2x spike that, on a full
    # month's export, pushed peak RSS toward Render's 512MB cap (OOM /
    # SIGSEGV). Popping each list keeps only one copy live at a time.
    import gc
    df = pd.DataFrame(index=range(len(data[keep_names[0]])) if keep_names else None)
    for n in keep_names:
        df[n] = data[n]
        data[n] = None
    del data
    gc.collect()
    return df


def parse_sessions(path: Path) -> pd.DataFrame:
    """Parse the export down to one row per inbound session (pre-tiering).

    Returns a DataFrame with columns:
      session_id, outcome, queue, is_spam, start_time, from_number, in_business_hours
    plus module-level counters attached as DataFrame.attrs.
    """
    # Memory-efficient read: a full month's Calls export is ~40MB / hundreds of
    # thousands of leg rows. pd.read_excel loads the WHOLE workbook (every
    # column, every row) into memory at once, and openpyxl's default mode keeps
    # the entire parsed sheet resident — together that blew past Render's 512MB
    # cap (OOM). Instead we stream the sheet in openpyxl read_only mode, keep
    # only the handful of columns we actually use, and drop non-inbound legs as
    # we go so they never accumulate.
    df = _read_calls_sheet_streaming(path)
    raw_inbound_legs = len(df)

    # Build derived columns from standalone Series (not df-views assigned back
    # into df) so pandas' copy-on-write machinery doesn't emit a stream of
    # ChainedAssignmentError FutureWarnings.
    handle_src = df["Handle Time"] if "Handle Time" in df.columns else pd.Series(0, index=df.index)
    df["_handle_seconds"] = handle_src.map(_to_seconds)
    df["_call_seconds"] = df["Call Length"].map(_to_seconds)
    if "Call Start Time" in df.columns:
        df["_start"] = df["Call Start Time"].map(_parse_start_time)
    else:
        df["_start"] = None
    df["_row_order"] = range(len(df))

    # ------------------------------------------------------------------
    # Collapse legs -> one row per session via VECTORIZED groupby.
    # (A per-group Python loop here is O(sessions) with heavy per-group
    # pandas ops and blew up memory/time on large exports — ~60k sessions
    # timed the worker out. Everything below is expressed as column ops.)
    # ------------------------------------------------------------------
    # df is already in export row order; preserve it for "first non-blank" picks.
    result_str = df["Result"].astype(str).str.strip()
    df["_ans_leg"] = (df["_handle_seconds"] > 0) | (result_str == OUTCOME_ANSWERED)
    df["_vmab"] = result_str == OUTCOME_VM_ABANDONED
    df["_vmmiss"] = result_str == OUTCOME_VM_MISSED
    df["_aband"] = result_str == OUTCOME_ABANDONED

    # Normalize Queue / From Number so blanks become NA and groupby.first()
    # naturally returns the first NON-blank value in row order.
    q = df["Queue"].astype("string").str.strip()
    df["_queue_norm"] = q.mask(q.isna() | (q == ""), other=pd.NA)
    if "From Number" in df.columns:
        fn = df["From Number"].astype("string").str.strip()
        df["_from_norm"] = fn.mask(fn.isna() | (fn == ""), other=pd.NA)
    else:
        df["_from_norm"] = pd.array([pd.NA] * len(df), dtype="string")

    gb = df.groupby("Session Id", sort=False)
    handle_max = gb["_handle_seconds"].max()
    call_max = gb["_call_seconds"].max()
    ans = gb["_ans_leg"].max()
    vmab = gb["_vmab"].max()
    vmmiss = gb["_vmmiss"].max()
    aband = gb["_aband"].max()
    queue_first = gb["_queue_norm"].first()      # first non-NA in row order
    from_first = gb["_from_norm"].first()
    start_first = gb["_start"].first() if "_start" in df.columns else None

    # Outcome by priority: Answered > VM/Abandoned > VM/Missed > Abandoned > Missed
    outcome = np.select(
        [ans.values, vmab.values, vmmiss.values, aband.values],
        [OUTCOME_ANSWERED, OUTCOME_VM_ABANDONED, OUTCOME_VM_MISSED, OUTCOME_ABANDONED],
        default=OUTCOME_MISSED,
    )

    start_vals = start_first.values if start_first is not None else np.array([None] * len(handle_max))

    sdf = pd.DataFrame({
        "session_id": handle_max.index.to_numpy(),
        "outcome": outcome,
        "queue": queue_first.fillna("Unknown").astype(str).to_numpy(),
        "handle_seconds": handle_max.astype(float).to_numpy(),
        "is_spam": (call_max <= SPAM_MAX_SECONDS).to_numpy(),
        "start_time": start_vals,
        "from_number": from_first.fillna("").astype(str).to_numpy(),
    })

    # Business-hours flag, vectorized from the (python datetime / None) starts.
    st = pd.to_datetime(pd.Series(start_vals), errors="coerce")
    in_bh = (
        st.notna()
        & st.dt.weekday.isin(BUSINESS_DAYS)
        & (st.dt.hour >= BUSINESS_START_HOUR)
        & (st.dt.hour < BUSINESS_END_HOUR)
    )
    sdf["in_business_hours"] = in_bh.to_numpy()

    sdf.attrs["raw_inbound_legs"] = raw_inbound_legs
    return sdf


def build_result(sdf: pd.DataFrame, queue_tiers: dict[str, dict],
                 queues_report: Optional[QueuesReport] = None) -> PipelineResult:
    """Apply tiering, spam filter, and compute all headline figures.

    queue_tiers: {queue_name: {"tier": "A".."D", "classification": "..."}}
    queues_report: optional parsed Queues report supplying real abandoned data.
    """
    raw_inbound_legs = sdf.attrs.get("raw_inbound_legs", len(sdf))
    inbound_sessions = len(sdf)
    phantom_legs_removed = raw_inbound_legs - inbound_sessions

    # Attach tier
    sdf = sdf.copy()
    sdf["tier"] = sdf["queue"].map(lambda q: queue_tiers.get(q, {}).get("tier", "C"))
    sdf["classification"] = sdf["queue"].map(lambda q: queue_tiers.get(q, {}).get("classification", ""))

    spam_mask = sdf["is_spam"]
    spam_sessions_removed = int(spam_mask.sum())
    clean = sdf[~spam_mask].copy()

    # Direct-dial calls (Tier D "Direct line") are excluded from the headline
    # universe below, but we tally them here for the secondary "and there's
    # more" callout. These are calls dialed to a personal extension, not the
    # main/intake queues, so a miss here is a weaker (existing-relationship)
    # signal than a missed main-queue call.
    _direct = clean[clean["queue"] == BA_QUEUE_DIRECT]
    direct_total = int(len(_direct))
    direct_missed = int((_direct["outcome"] != OUTCOME_ANSWERED).sum())
    direct_abandoned = int(_direct["outcome"].isin(
        [OUTCOME_ABANDONED, OUTCOME_VM_ABANDONED]).sum())

    # Headline universe: A+B+C queues only (exclude Tier D back-office).
    #
    # Also exclude un-queued sessions (blank Queue -> "Unknown"). These are
    # direct dials to a personal extension, not calls that entered a customer
    # call queue. Including them (they previously defaulted to Tier C) inflated
    # the missed-call rate: a direct call to someone's desk/cell that rings out
    # or hits personal voicemail is not a missed revenue-line queue call, and it
    # never appears in the RingCentral Queues Performance report — so the deck's
    # number couldn't reconcile against the report an AE would verify it with.
    # Restricting to real queues makes the miss rate defensible and verifiable.
    universe = clean[clean["tier"].isin(["A", "B", "C"])
                     & (clean["queue"] != "Unknown")].copy()

    if len(universe) == 0:
        n_clean = len(clean)
        if n_clean == 0:
            raise ValueError(
                "No usable inbound calls were found in this export after de-duplication "
                "and spam filtering. Make sure you exported the Calls detail view with "
                "inbound calls in the selected date range."
            )
        raise ValueError(
            f"Found {n_clean} inbound call(s), but none of them entered a customer-facing "
            "call queue (they were back-office queues or direct dials to personal "
            "extensions, both excluded from the revenue analysis). Re-export including "
            "your sales / retail / branch call queues, or pick a date range with "
            "customer-facing queue volume."
        )

    def count(df, outcome):
        return int((df["outcome"] == outcome).sum())

    answered = count(universe, OUTCOME_ANSWERED)
    vm_abandoned = count(universe, OUTCOME_VM_ABANDONED)
    vm_missed = count(universe, OUTCOME_VM_MISSED)
    abandoned = count(universe, OUTCOME_ABANDONED)
    missed = count(universe, OUTCOME_MISSED)
    universe_sessions = len(universe)

    total_missed = vm_abandoned + vm_missed + abandoned + missed
    missed_df = universe[universe["outcome"] != OUTCOME_ANSWERED]

    # Answered calls with under 60s talk time (AIR-solvable "basic" calls)
    answered_df = universe[universe["outcome"] == OUTCOME_ANSWERED]
    if "handle_seconds" in answered_df.columns:
        answered_under_60 = int(((answered_df["handle_seconds"] > 0)
                                 & (answered_df["handle_seconds"] < 60)).sum())
    else:
        answered_under_60 = 0

    # Average talk time of answered calls (for AIR usage-cost model)
    if "handle_seconds" in answered_df.columns and len(answered_df):
        hs = answered_df["handle_seconds"]
        hs = hs[hs > 0]
        avg_answered_minutes = float(hs.mean() / 60.0) if len(hs) else 3.0
    else:
        avg_answered_minutes = 3.0

    # Business hours miss %
    bh_miss = int(missed_df["in_business_hours"].sum())
    business_hours_miss_pct = bh_miss / total_missed if total_missed else 0.0

    # Repeat callers: distinct From Number with 2+ unanswered sessions
    repeat_callers = 0
    if "from_number" in missed_df.columns:
        counts = missed_df[missed_df["from_number"] != ""]["from_number"].value_counts()
        repeat_callers = int((counts >= 2).sum())

    # Hourly distribution of misses (business hours window)
    hourly_missed = {h: 0 for h in range(BUSINESS_START_HOUR, BUSINESS_END_HOUR)}
    for st in missed_df["start_time"].dropna():
        if st.hour in hourly_missed:
            hourly_missed[st.hour] += 1

    # Miss-rate-by-hour across all 24 hours + weekend / midday (reference deck slide 3)
    # NB: build _hour/_weekday via list comprehensions so the columns are always
    # integer-typed — `Series.apply` on an *empty* datetime64 column would otherwise
    # return an empty datetime64 series and blow up on `>= 0` comparisons.
    uni = universe.copy()

    def _ok_ts(s):
        return s is not None and not pd.isna(s)

    uni["_hour"] = [s.hour if _ok_ts(s) else -1 for s in uni["start_time"]]
    uni["_weekday"] = [s.weekday() if _ok_ts(s) else -1 for s in uni["start_time"]]
    uni["_missed"] = uni["outcome"] != OUTCOME_ANSWERED
    uni_ts = uni[uni["_hour"] >= 0]

    hourly_inbound = {h: 0 for h in range(24)}
    hourly_miss_rate = {h: 0.0 for h in range(24)}
    for h in range(24):
        sub = uni_ts[uni_ts["_hour"] == h]
        hourly_inbound[h] = len(sub)
        hourly_miss_rate[h] = float(sub["_missed"].mean()) if len(sub) else 0.0

    def _rate_range(hours):
        rates = [hourly_miss_rate[h] for h in hours if hourly_inbound[h] > 0]
        return (min(rates), max(rates)) if rates else (0.0, 0.0)

    after_hours = list(range(18, 24)) + list(range(0, 6))   # 6pm-6am
    after_hours_miss_lo, after_hours_miss_hi = _rate_range(after_hours)
    ah_df = uni_ts[uni_ts["_hour"].isin(after_hours)]
    after_hours_miss_rate = float(ah_df["_missed"].mean()) if len(ah_df) else 0.0

    sat = uni_ts[uni_ts["_weekday"] == 5]
    sun = uni_ts[uni_ts["_weekday"] == 6]
    wk_rates = [df["_missed"].mean() for df in (sat, sun) if len(df)]
    weekend_miss_lo = float(min(wk_rates)) if wk_rates else 0.0
    weekend_miss_hi = float(max(wk_rates)) if wk_rates else 0.0
    wk_df = uni_ts[uni_ts["_weekday"].isin([5, 6])]
    weekend_miss_rate = float(wk_df["_missed"].mean()) if len(wk_df) else 0.0

    midday = uni_ts[uni_ts["_hour"].isin([11, 12, 13])]
    midday_miss_rate = float(midday["_missed"].mean()) if len(midday) else 0.0

    # Period span
    starts = [s for s in universe["start_time"].dropna()]
    if starts:
        dmin, dmax = min(starts).date(), max(starts).date()
        days_in_period = (dmax - dmin).days + 1
        reporting_period = f"{dmin.strftime('%b %-d')}–{dmax.strftime('%-d, %Y')}"
    else:
        days_in_period = 30
        reporting_period = ""

    # Per-queue stats (A+B+C)
    queue_stats: dict[str, QueueStats] = {}
    for _, row in universe.iterrows():
        q = row["queue"]
        if q not in queue_stats:
            qs = QueueStats(name=q)
            qs.tier = row["tier"]
            qs.classification = row["classification"]
            queue_stats[q] = qs
        qs = queue_stats[q]
        qs.inbound += 1
        o = row["outcome"]
        if o == OUTCOME_ANSWERED:
            qs.answered += 1
            hs = row.get("handle_seconds", 0) or 0
            if 0 < hs < 60:
                qs.answered_under_60 += 1
        elif o == OUTCOME_VM_ABANDONED:
            qs.vm_abandoned += 1
        elif o == OUTCOME_VM_MISSED:
            qs.vm_missed += 1
        elif o == OUTCOME_ABANDONED:
            qs.abandoned += 1
        else:
            qs.missed += 1

    # Tier the Queues-report queues (back-office excluded from the abandoned story)
    if queues_report:
        for qa in queues_report.queues:
            meta = queue_tiers.get(qa.name, {})
            qa.tier = (meta.get("tier") or "C")
            qa.classification = meta.get("classification") or ""

    computed = answered + total_missed
    recon_ok = computed == universe_sessions
    recon_note = (f"OK: {computed} outcomes == {universe_sessions} universe sessions"
                  if recon_ok else
                  f"MISMATCH: {computed} != {universe_sessions}")

    return PipelineResult(
        raw_inbound_legs=raw_inbound_legs,
        inbound_sessions=inbound_sessions,
        phantom_legs_removed=phantom_legs_removed,
        spam_sessions_removed=spam_sessions_removed,
        universe_sessions=universe_sessions,
        answered=answered,
        vm_abandoned=vm_abandoned,
        vm_missed=vm_missed,
        abandoned=abandoned,
        missed=missed,
        business_hours_miss_pct=business_hours_miss_pct,
        repeat_callers=repeat_callers,
        hourly_missed=hourly_missed,
        days_in_period=days_in_period,
        answered_under_60=answered_under_60,
        hourly_miss_rate=hourly_miss_rate,
        hourly_inbound=hourly_inbound,
        after_hours_miss_lo=after_hours_miss_lo,
        after_hours_miss_hi=after_hours_miss_hi,
        after_hours_miss_rate=after_hours_miss_rate,
        weekend_miss_lo=weekend_miss_lo,
        weekend_miss_hi=weekend_miss_hi,
        weekend_miss_rate=weekend_miss_rate,
        midday_miss_rate=midday_miss_rate,
        avg_answered_minutes=avg_answered_minutes,
        direct_total=direct_total,
        direct_missed=direct_missed,
        direct_abandoned=direct_abandoned,
        queue_stats=queue_stats,
        queues_report=queues_report,
        reporting_period=reporting_period,
        reconciliation_ok=recon_ok,
        reconciliation_note=recon_note,
        sessions_df=sdf,
    )


def distinct_queues(sdf: pd.DataFrame) -> list[str]:
    return sorted(q for q in sdf["queue"].dropna().unique() if q and q != "Unknown")


def _to_int(val) -> int:
    if pd.isna(val):
        return 0
    try:
        return int(round(float(str(val).strip().replace(",", ""))))
    except (ValueError, TypeError):
        return 0


def _to_pct(val) -> float:
    """Parse a percentage cell that may be '52.3', '52.3%' or a 0-1 fraction."""
    if pd.isna(val):
        return 0.0
    s = str(val).strip().replace("%", "")
    try:
        f = float(s)
    except ValueError:
        return 0.0
    return f / 100.0 if f > 1.0 else f


def parse_queues_report(path: Path) -> QueuesReport:
    """Parse the RingCentral *Queues* performance report (the 2nd upload).

    Expects the workbook to contain a ``KPIs`` sheet (account totals) and a
    ``Queues`` sheet (per-queue rows). Raises ValueError with a clear message
    if the file looks like the Calls export or is otherwise unrecognized.
    """
    xl = pd.ExcelFile(path)
    names_lower = {n.strip().lower(): n for n in xl.sheet_names}

    kpi_sheet = names_lower.get("kpis")
    queues_sheet = names_lower.get("queues")
    if not kpi_sheet or not queues_sheet:
        raise ValueError(
            "This doesn't look like a RingCentral Queues report. It should have "
            "'KPIs' and 'Queues' tabs (Performance Reports → Queues → Download → Excel). "
            f"Found sheets: {xl.sheet_names}."
        )

    kdf = pd.read_excel(path, sheet_name=kpi_sheet, dtype=str)
    kdf.columns = [c.strip() for c in kdf.columns]
    if "# Abandoned" not in kdf.columns or len(kdf) == 0:
        raise ValueError(
            "The Queues report's KPIs tab is missing the '# Abandoned' column. "
            "Re-download the Queues performance report from analytics.ringcentral.com."
        )
    k = kdf.iloc[0]

    report = QueuesReport(
        inbound=_to_int(k.get("# Inbound")),
        answered=_to_int(k.get("# Answered")),
        abandoned=_to_int(k.get("# Abandoned")),
        voicemail=_to_int(k.get("# Voicemail")),
        avg_speed_answer=str(k.get("Avg. Speed of Answer") or "").strip(),
        avg_wait=str(k.get("Avg. Wait Time") or "").strip(),
        longest_wait=str(k.get("Longest Wait Time") or "").strip(),
        sla_pct=_to_pct(k.get("% SLA")),
    )

    qdf = pd.read_excel(path, sheet_name=queues_sheet, dtype=str)
    qdf.columns = [c.strip() for c in qdf.columns]
    for _, row in qdf.iterrows():
        name = str(row.get("Name") or "").strip()
        if not name:
            continue
        inbound = _to_int(row.get("# Inbound"))
        if inbound == 0 and _to_int(row.get("# Abandoned")) == 0:
            continue  # skip empty/idle queues
        report.queues.append(QueueAbandon(
            name=name,
            ext=str(row.get("Ext") or "").strip(),
            inbound=inbound,
            answered=_to_int(row.get("# Answered")),
            abandoned=_to_int(row.get("# Abandoned")),
        ))
    return report


def queues_report_queue_names(report: Optional[QueuesReport]) -> list[str]:
    """Queue names from the Queues report, for inclusion in tiering."""
    if not report:
        return []
    return [q.name for q in report.queues]


# ===========================================================================
# Business Analytics single-report pipeline (new foundation, 2026-07)
# ---------------------------------------------------------------------------
# RingCentral is deprecating Performance Reports. The new source of truth is the
# Business Analytics "Call Records" widget, exported as ONE Excel file. It is
# per-call (not per-leg), and — crucially — RingCentral now stamps each call's
# outcome directly in a `Result` column, so we no longer re-derive missed/
# abandoned/voicemail from durations or stitch a second Queues report. We simply
# trust RingCentral's own labels, which means every headline number reconciles
# 1:1 with the Business Analytics dashboard KPI tiles the AE can see on screen.
#
# File shape:
#   rows 0-5  : a metadata block (Widget Name, Start/End Date, TimeZone, filter)
#   row  6    : the real column header
#   row  7+   : one row per call
# Header columns:
#   From Name, From Number, To Name, To Number, Date-Time, Length, Direction,
#   Call Type, Call Response, Result, Ringing, IVR Prompt, Live Talk, Hold,
#   Park, Transfer, VM Greeting, VoiceMail, Setup, Forwarding, Origin,
#   Call Id, Hop Id
#
# There is no Queue dimension in this widget, so all inbound is reported under a
# single synthetic "Direct line" queue (Tier C) — this keeps the existing
# build_result math and every deck slide working unchanged.
# ===========================================================================

BA_SYNTHETIC_QUEUE = "Direct line"

# Serviced-vs-unserviced framing. RingCentral's `Call Type` distinguishes calls
# that entered a shared call queue (main/intake lines — potential NEW customers)
# from calls dialed straight to one person's extension (existing relationships,
# callbacks). Only the former belong in the headline missed-revenue case; the
# latter are broken out as a smaller secondary figure. See parse_business_analytics.
BA_QUEUE_INTAKE = "Main call queues"   # Call Type = Queue Calls / Overflow / Transferred
BA_QUEUE_DIRECT = "Direct line"        # Call Type = Inbound Direct (personal extension)
# Call Type values that mean "entered a shared/main call queue".
_BA_QUEUE_TYPES = {"queue calls", "overflow calls", "transferred calls", "park retrievals"}

# RingCentral Business-Analytics `Result` values -> our internal outcomes.
# Business Analytics has no separate "abandoned-to-voicemail" state, so
# vm_abandoned is always 0 on this source.
_BA_RESULT_MAP = {
    "completed": OUTCOME_ANSWERED,
    "transferred": OUTCOME_ANSWERED,   # handled — transferred to another party
    "missed with vm": OUTCOME_VM_MISSED,
    "missed without vm": OUTCOME_MISSED,
    "abandoned": OUTCOME_ABANDONED,
}

_BA_HEADER_MARKERS = {"from number", "result", "direction", "date-time"}


def _ba_seconds(val) -> float:
    """Business Analytics duration cells come through openpyxl as
    datetime.timedelta (Excel time-of-day). Convert any of timedelta / number /
    'HH:MM:SS' string to float seconds."""
    if val is None:
        return 0.0
    if isinstance(val, dt.timedelta):
        return val.total_seconds()
    if isinstance(val, (int, float)):
        # Excel serial time-of-day is a day fraction; but analytics exports come
        # as timedelta, so a bare number here is already seconds.
        return float(val)
    return _to_seconds(val)


def _ba_parse_datetime(val):
    """Parse a Business Analytics 'Date-Time' cell (e.g. '07/20/2026 9:52:58 PM'
    or a real datetime) into a python datetime, or None."""
    if val is None:
        return None
    if isinstance(val, dt.datetime):
        return val
    return _parse_start_time(val)


def parse_business_analytics(path: Path) -> pd.DataFrame:
    """Parse a RingCentral Business Analytics 'Call Records' export into one row
    per inbound external session, in the SAME schema build_result() consumes:
      session_id, outcome, queue, handle_seconds, is_spam, start_time,
      from_number, in_business_hours   (+ attrs['raw_inbound_legs']).

    Trusts RingCentral's own `Result` label for the outcome. Filters to
    Direction=Inbound and Origin=External, de-duplicates transfer hops by
    Call Id (answered wins), and drops unclassifiable 'Other' results as spam so
    the universe matches the dashboard's classified inbound count.
    """
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        # This widget exports a single sheet ("Calls_Records"); take the first.
        ws = wb[wb.sheetnames[0]]

        header = None
        header_i = None
        col = {}
        rows_iter = ws.iter_rows(values_only=True)
        # Locate the real header row (skip the metadata block at the top).
        for i, row in enumerate(rows_iter):
            cells = [str(c).strip().lower() if c is not None else "" for c in row]
            if _BA_HEADER_MARKERS.issubset(set(cells)):
                header = [str(c).strip() if c is not None else "" for c in row]
                header_i = i
                col = {h.lower(): j for j, h in enumerate(header) if h}
                break
        if header is None:
            raise ValueError(
                "This doesn't look like a RingCentral Business Analytics 'Call "
                "Records' export. Expected a header row with From Number / "
                "Direction / Result / Date-Time columns. In Analytics → Business "
                "Analytics, open the Call Records dashboard, add the Call "
                "Segment / Origin / Call ID / Hop ID / Recording columns, Save, "
                "then Download."
            )

        def gi(name):
            return col.get(name.lower())

        i_dir, i_origin, i_result = gi("Direction"), gi("Origin"), gi("Result")
        i_from, i_dt = gi("From Number"), gi("Date-Time")
        i_talk = gi("Live Talk")
        i_len = gi("Length")
        i_callid, i_type = gi("Call Id"), gi("Call Type")

        records = []
        raw_inbound_legs = 0
        for row in rows_iter:  # continues AFTER the header row
            if i_dir is None or i_dir >= len(row):
                continue
            direction = str(row[i_dir] or "").strip().lower()
            if direction != "inbound":
                continue
            origin = str(row[i_origin] or "").strip().lower() if i_origin is not None else "external"
            if origin and origin != "external":
                continue  # drop internal extension-to-extension traffic
            raw_inbound_legs += 1
            result = str(row[i_result] or "").strip()
            call_type = str(row[i_type] or "").strip() if i_type is not None else ""
            records.append({
                "call_id": str(row[i_callid] or "").strip() if i_callid is not None else "",
                "result": result,
                "call_type": call_type,
                "from_number": str(row[i_from] or "").strip() if i_from is not None else "",
                "start_time": _ba_parse_datetime(row[i_dt] if i_dt is not None else None),
                "talk_seconds": _ba_seconds(row[i_talk] if i_talk is not None else None),
                "len_seconds": _ba_seconds(row[i_len] if i_len is not None else None),
            })
    finally:
        wb.close()

    if not records:
        raise ValueError(
            "No inbound calls were found in this Business Analytics export. "
            "Make sure the dashboard date range covers a period with inbound "
            "call volume before downloading."
        )

    df = pd.DataFrame.from_records(records)

    # Map RingCentral's Result label -> outcome. Unknown/'Other' -> spam (excluded).
    def map_outcome(res: str) -> str:
        return _BA_RESULT_MAP.get(res.strip().lower(), "")

    df["outcome"] = df["result"].map(map_outcome)
    df["is_spam"] = df["outcome"] == ""            # unclassifiable ('Other' etc.)
    df.loc[df["is_spam"], "outcome"] = OUTCOME_MISSED  # placeholder; filtered out

    # Classify each leg as a queue (main/intake) call vs a direct-dial call.
    # A physical call is a "queue" call if ANY of its legs entered a call queue.
    df["_is_queue"] = (df["call_type"].fillna("").str.strip().str.lower()
                       .isin(_BA_QUEUE_TYPES))

    # De-duplicate transfer hops: one physical call can appear as several rows
    # sharing a Call Id. Collapse to one session, answered-priority.
    _priority = {OUTCOME_ANSWERED: 5, OUTCOME_VM_ABANDONED: 4, OUTCOME_VM_MISSED: 3,
                 OUTCOME_ABANDONED: 2, OUTCOME_MISSED: 1}
    df["_rank"] = df["outcome"].map(lambda o: _priority.get(o, 0))
    has_id = df["call_id"].astype(str).str.len() > 0
    # Per-call queue flag = did ANY leg of this Call Id enter a queue?
    qmap = df[has_id].groupby("call_id")["_is_queue"].any() if has_id.any() else {}
    keyed, unkeyed = df[has_id].copy(), df[~has_id].copy()
    if len(keyed):
        keyed = (keyed.sort_values("_rank", ascending=False)
                      .groupby("call_id", sort=False, as_index=False).first())
        keyed["_is_queue"] = keyed["call_id"].map(qmap).fillna(False)
    sdf_src = pd.concat([keyed, unkeyed], ignore_index=True)

    handle = np.where(sdf_src["outcome"] == OUTCOME_ANSWERED,
                      sdf_src["talk_seconds"].astype(float), 0.0)
    starts = list(sdf_src["start_time"])
    st = pd.to_datetime(pd.Series(starts), errors="coerce")
    in_bh = (
        st.notna()
        & st.dt.weekday.isin(BUSINESS_DAYS)
        & (st.dt.hour >= BUSINESS_START_HOUR)
        & (st.dt.hour < BUSINESS_END_HOUR)
    )

    sdf = pd.DataFrame({
        "session_id": (sdf_src["call_id"].where(sdf_src["call_id"].astype(str).str.len() > 0,
                       [f"row-{i}" for i in range(len(sdf_src))]).to_numpy()),
        "outcome": sdf_src["outcome"].to_numpy(),
        "queue": np.where(sdf_src["_is_queue"].to_numpy(),
                          BA_QUEUE_INTAKE, BA_QUEUE_DIRECT),
        "handle_seconds": handle.astype(float),
        "is_spam": sdf_src["is_spam"].to_numpy(),
        "start_time": starts,
        "from_number": sdf_src["from_number"].fillna("").astype(str).to_numpy(),
    })
    sdf["in_business_hours"] = in_bh.to_numpy()
    sdf.attrs["raw_inbound_legs"] = raw_inbound_legs

    # Queue-segment completeness. The whole business case rests on missed calls
    # to the main / intake CALL QUEUES. The default Call Records export omits the
    # queue-call segments, leaving a file that is almost all direct-dial rows —
    # which silently produces a weak deck that measures the wrong thing. Detect
    # that here (on non-spam calls) and stash the counts + a warning message so
    # the app can block the run with an actionable, specific error.
    _live = sdf[~sdf["is_spam"]]
    q_calls = int((_live["queue"] == BA_QUEUE_INTAKE).sum())
    d_calls = int((_live["queue"] == BA_QUEUE_DIRECT).sum())
    inbound_calls = int(len(_live))
    q_share = (q_calls / inbound_calls) if inbound_calls else 0.0
    sdf.attrs["queue_calls"] = q_calls
    sdf.attrs["direct_calls"] = d_calls
    sdf.attrs["queue_call_share"] = q_share
    # Misconfigured export = queue calls are both a tiny share AND a tiny count.
    # (A genuinely small firm with real queues still shows a healthy share, so
    # this won't false-positive on low volume.)
    sdf.attrs["queue_segments_missing"] = (q_share < 0.15 and q_calls < 500)
    return sdf


# Threshold constants surfaced for the app's pre-flight check / messaging.
QUEUE_SEGMENT_MIN_SHARE = 0.15
QUEUE_SEGMENT_MIN_CALLS = 500


def queue_segment_error(sdf: pd.DataFrame) -> Optional[str]:
    """Return an actionable error message if the export is missing its main /
    intake queue-call segments, else None. The business case is built on missed
    MAIN-QUEUE calls; without the queue segments the file is almost all
    direct-dial traffic and the deck would understate the real opportunity."""
    if not sdf.attrs.get("queue_segments_missing"):
        return None
    q = sdf.attrs.get("queue_calls", 0)
    d = sdf.attrs.get("direct_calls", 0)
    return (
        f"This export is missing its main call-queue data, so a business case can't "
        f"be built from it. We found only {q:,} calls to your main / intake queues "
        f"(vs {d:,} direct-dial calls to individual extensions). The AI Business Case "
        f"is built on missed calls to your MAIN queues — the potential new customers — "
        f"and the default Call Records export leaves the queue-call segments out.\n\n"
        f"Re-export from RingCentral Analytics → Business Analytics → Call Records with "
        f"the QUEUE CALL segments included (not just the default view). A correct export "
        f"is large (tens of MB) and its Call Type column shows thousands of 'Queue Calls' "
        f"rows. Then upload it again."
    )


def ba_queue_tiers() -> dict[str, dict]:
    """Tiering map for the two Business-Analytics call buckets, split by Call Type.

    Main call queues (intake / shared lines) are the potential-new-customer
    traffic and form the headline universe (Tier A). Direct-dial calls to a
    personal extension are existing relationships/callbacks; they are Tier D so
    they drop out of the headline missed-revenue number and are surfaced only as
    a smaller secondary figure (see PipelineResult.direct_* fields)."""
    return {
        BA_QUEUE_INTAKE: {"tier": "A", "classification": "Main / intake call queue"},
        BA_QUEUE_DIRECT: {"tier": "D", "classification": "Direct dial to a person"},
    }
